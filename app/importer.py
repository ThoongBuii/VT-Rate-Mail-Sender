from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .models import AgencyMail, MailStatus

COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "agency_company": (
        "agency company",
        "agency_company",
        "agency",
        "company",
        "công ty",
        "cong ty",
    ),
    "account_name": (
        "account name",
        "account_name",
        "name",
        "tên",
        "ten",
        "người nhận",
        "nguoi nhan",
    ),
    "account_mail": (
        "account mail",
        "account_mail",
        "email",
        "mail",
        "to",
        "email nhận",
        "email nhan",
    ),
    "mail_cc": (
        "mail cc",
        "mail_cc",
        "cc",
        "cc mail",
        "email cc",
    ),
    # Optional legacy columns (ignored if compose UI is used)
    "subject": ("subject", "tiêu đề", "tieu de"),
    "attachment": (
        "attachment",
        "file đính kèm",
        "file dinh kem",
        "đính kèm",
        "dinh kem",
        "attach",
    ),
    "template_mail": (
        "template mail",
        "template_mail",
        "template",
        "body",
        "nội dung",
        "noi dung",
        "content",
    ),
}


def _normalize_header(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _map_headers(headers: Iterable[str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    normalized = [_normalize_header(h) for h in headers]
    for field, aliases in COLUMN_ALIASES.items():
        for idx, header in enumerate(normalized):
            if header in aliases:
                mapping[field] = idx
                break
    return mapping


def _cell(row: list, idx: int | None) -> str:
    if idx is None or idx >= len(row):
        return ""
    value = row[idx]
    if value is None:
        return ""
    return str(value).strip()


def _rows_to_mails(headers: list[str], data_rows: list[list], start_row: int = 2) -> list[AgencyMail]:
    mapping = _map_headers(headers)
    if "account_mail" not in mapping:
        raise ValueError(
            "Không tìm thấy cột Account Mail.\n"
            "Excel chỉ cần: Agency Company, Account Name, Account Mail, Mail cc"
        )
    result: list[AgencyMail] = []
    for i, values in enumerate(data_rows, start=start_row):
        if all(v is None or str(v).strip() == "" for v in values):
            continue
        mail = AgencyMail(
            agency_company=_cell(values, mapping.get("agency_company")),
            account_name=_cell(values, mapping.get("account_name")),
            account_mail=_cell(values, mapping.get("account_mail")),
            mail_cc=_cell(values, mapping.get("mail_cc")),
            # Legacy optional — UI compose sẽ ghi đè
            subject=_cell(values, mapping.get("subject")),
            attachment=_cell(values, mapping.get("attachment")),
            template_mail=_cell(values, mapping.get("template_mail")),
            status=MailStatus.PENDING,
            row_index=i,
        )
        result.append(mail)
    return result


def import_excel(path: str | Path) -> list[AgencyMail]:
    path = Path(path)
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(c or "") for c in rows[0]]
    return _rows_to_mails(headers, [list(r) for r in rows[1:]])


def import_csv(path: str | Path) -> list[AgencyMail]:
    path = Path(path)
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    if not rows:
        return []
    return _rows_to_mails(rows[0], rows[1:])


def import_agencies(path: str | Path) -> list[AgencyMail]:
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return import_excel(path)
    if suffix == ".csv":
        return import_csv(path)
    raise ValueError("Chỉ hỗ trợ file .xlsx hoặc .csv")


def export_sample_excel(path: str | Path) -> Path:
    """Excel chỉ 4 cột danh bạ — Subject/Attachment/Template soạn trên app."""
    path = Path(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Agencies"
    headers = ["Agency Company", "Account Name", "Account Mail", "Mail cc"]
    ws.append(headers)
    ws.append(["VT Internal TEST", "Amber", "your.email@company.com", ""])
    ws.append(["ABC Logistics Co., Ltd", "Nguyen Van A", "vana@abclogistics.com", "ops@abclogistics.com"])
    ws.append(["Global Freight Partners", "Tran Thi B", "thib@gfp.com", ""])

    guide = wb.create_sheet("HUONG_DAN")
    guide.append(["Hướng dẫn"])
    guide.append(["1", "Excel chỉ cần 4 cột: Company / Account Name / Account Mail / Mail cc"])
    guide.append(["2", "Subject, Attachment, Template Mail — soạn trên giao diện app"])
    guide.append(["3", "Sửa Account Mail dòng test thành email của bạn trước khi gửi thử"])

    fill = PatternFill("solid", fgColor="0F2C4C")
    font = Font(color="FFFFFF", bold=True)
    thin = Border(
        left=Side(style="thin", color="D0D7DE"),
        right=Side(style="thin", color="D0D7DE"),
        top=Side(style="thin", color="D0D7DE"),
        bottom=Side(style="thin", color="D0D7DE"),
    )
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center")
        cell.border = thin
    ws["C2"].fill = PatternFill("solid", fgColor="FFF3CD")
    widths = [28, 18, 34, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    wb.save(path)
    return path
